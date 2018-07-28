# Copyright (c) 2016, Oracle and/or its affiliates. All rights reserved.
#
#
# author:
#
# modified by: Kiran K Sanghisetti
#

import os, sys
import openpyxl
import ipcalc
import optparse
import re

try:
    import json
except ImportError:
    import simplejson as json

from subprocess import PIPE, Popen

def run_command(command):
    process = Popen(
        args=command,
        stdout=PIPE,
        stderr=PIPE,
        shell=True
    )

    output, error = process.communicate()
    return (output.strip() , error.strip())

def remove_file(filename):
    try:
        os.remove(filename)
    except OSError:
        pass

    return

def append_to_file(xls_file_path, string):
    try:
        file_handle = open(xls_file_path, 'a')
        file_handle.write(string)
        file_handle.close()
    except IOError:
        pass
    
class SimpleCell(object):
    def __init__(self, row=0, col=0):
        self.row = row
        self.col = col

class XLSheet(object):
    
    def __verify_ip_value(self, value):
        try:
            ipcalc.Network(value)
            ip = value
        except ValueError:
            ip = self.__get_ip(value)
        except:
            raise ValueError('wrong value - %s' %(value))
        
        return ip
    
    def __init__(self, sheet, pod_dict, txt_file_name):
        self.sheet = sheet
        self.hostname = SimpleCell(0,0)
        self.device = SimpleCell(0,0)
        self.interface = SimpleCell(0,0)
        self.subnet = SimpleCell(0,0)
        self.vlan = SimpleCell(0,0)
        
        self.bdcs_zone_id = pod_dict['bdcs_zone_id']
        self.internal_dns = pod_dict['internal_dns']
        self.internal_dns_domain = pod_dict['internal_dns_domain']
        self.external_dns_domain = pod_dict['external_dns_domain']
        self.ntp = pod_dict['ntp']
        self.asr = pod_dict['asr']
        self.nfs_hdfs_retention = pod_dict['nfs_hdfs_retention']
        self.nfs_staging = pod_dict['nfs_staging']
        self.bastion_ip = pod_dict['bastion_ip']
        self.bdcs_url = pod_dict['bdcs_url']
        self.bdcs_ip = pod_dict['bdcs_ip']
        self.em_manager_ip = pod_dict['em_manager_ip']
        
        self.txt_file_name = txt_file_name.lower() + '.txt'
        remove_file(self.txt_file_name)
        print '\nGenerating txt file - %s' %(self.txt_file_name)
    
        for row in self.sheet.rows:
            for cell in row:
                try:
                    cell_value = cell.value
                    cell_value = cell_value.lower()

                    if cell_value == 'HOSTNAME'.lower():
                        self.hostname.row = cell.row
                        self.hostname.col = cell.column
                    elif cell_value == 'DEVICE'.lower():
                        self.device.row = cell.row
                        self.device.col = cell.column
                    elif cell_value == 'Subnet Details'.lower():
                        self.subnet.row = cell.row
                        self.subnet.col = cell.column
                    elif cell_value == 'VLAN'.lower():
                        self.vlan.row = cell.row
                        self.vlan.col = cell.column
                    elif cell_value == 'INTERFACE CATEGORY'.lower():
                        self.interface.row = cell.row
                        self.interface.col = cell.column
                    elif self.hostname.row > 0 and self.subnet.row > 0 and self.vlan.row > 0:
                        break
                except:
                    pass
    
    def create_text_file(self):
	#
	# function to create rack text files
	#
	# args: none
	#
	# depends: 
	#	variables: 
	#	functions: hardcoded_info() get_nw_info() get_hostnames()
	#
	# returns: none

        self.hardcoded_info()
        self.get_nw_info()
        self.get_hostnames()


    def __append_to_file(self, msg):
        append_to_file(self.txt_file_name, '\n'+msg)

    
    def hardcoded_info(self):
	#
	# function to read and update json input data into rack text files

	# args: none
	#
	# depends: 
	#	variables: bdcs_zone_id internal_dns internal_dns_domain
	#	   	   external_dns_domain ntp asr nfs_hdfs_retention
	#		   nfs_staging bastion_ip bdcs_url bdcs_ip em_manager_ip
	#
	#	functions: __append_to_file
	#
	# returns: none

        self.__append_to_file('BDCS Zone ID:           %s' %(self.bdcs_zone_id))
        self.__append_to_file('DNS internal:           %s' %(self.internal_dns))
        self.__append_to_file('Internal DNS domain:    %s' %(self.internal_dns_domain))
        self.__append_to_file('External DNS domain:    %s' %(self.external_dns_domain))
        self.__append_to_file('NTP:                    %s' %(self.ntp))
        self.__append_to_file('ASR:                    %s' %(self.asr))
        self.__append_to_file('NFS HDFS Retention:     %s' %(self.nfs_hdfs_retention))
        self.__append_to_file('NFS Staging :           %s' %(self.nfs_staging))
        self.__append_to_file('Bastion Ip:             %s' %(self.bastion_ip))
        self.__append_to_file('Bdcs Url:               %s' %(self.bdcs_url))
        self.__append_to_file('Bdcs Ip:                %s' %(self.bdcs_ip))
        self.__append_to_file('Em Manager Ip:          %s' %(self.em_manager_ip))
        
    def __get_rows(self, col, start, end):
        value_list = []
        for rowNum in range(start, end):  # skip the first row
            try:
                value = self.sheet.cell(row=rowNum, column=col).value
                if value:
                    value_list.append(value)
            except:
                pass
        return value_list
    
    def __get_ip(self, hostname_val):
	#
	# function to resolve an hostname to IP
	#
	# args: hostname_val
	#
	# depends: 
	#	variables: internal_dns_domain
	#	functions: run_command()
	#
	# returns: string 

        output, error = run_command('host %s.%s' %(hostname_val,self.internal_dns_domain))
        return output.split(" ")[-1]

#    def __build_dev_list(self, col, rowNum):
#	device = self.sheet.cell(column=col, row=rowNum).value
#	hostname = self.sheet.cell(column=col+1, row=rowNum).value
#
#	if 'switch' in device.lower():
#	    self.dict['switch'].append(hostname)
#       elif 'pdu' in device.lower():
#	    self.dict['pdu'].append(hostname)
#        elif 'cisco' in device.lower():
#	    self.dict['cisco'].append(hostname)
#        elif 'bda server' in device.lower():
#	    self.dict['dom0'].append(hostname)
#	    self.dict['ilom'].append(hostname + '-ilom')
    
    def get_hostnames(self):
	#
	# function to grab all device names (dom0, switches, pdu)
	# from RE sheet

	# args: none
	#
	# depends: 
	#	variables: self.hostname.col self.hostname.row self.device.col
	#	   	   self.device.row
	#	functions: 
	#
	# returns: none

        #hostnames = self.__get_rows(self.hostname.col, self.hostname.row+1, self.hostname.row+55)
        #devices = self.__get_rows(self.device.col, self.device.row+1, self.device.row+55)

	dict = {}
	dict['dom0'] = []
	dict['ilom'] = []
	dict['switch'] = []
	dict['pdu'] = []
	dict['cisco'] = []

	col = self.hostname.col-1
	rowNum = self.hostname.row

	if self.sheet.cell(column=col, row=rowNum).value == "DEVICE":
	    rowNum = rowNum + 1
	    nullcount = 0

	    while True:
	        if self.sheet.cell(column=col, row=rowNum).value == None:
	            if nullcount > 2:
		        break

		    nullcount = nullcount + 1    
	            rowNum = rowNum + 1
		    continue
	        else:
		    nullcount = 0    
		    if self.sheet.cell(column=col+1, row=rowNum).value == None:
		        raise ValueError("Null value found in place of hostname")
		    else:
			#self.__build_dev_list(col, rowNum)
	                #rowNum = rowNum + 1
			###############
			device = self.sheet.cell(column=col, row=rowNum).value
			hostname = self.sheet.cell(column=col+1, row=rowNum).value

			if 'switch' in device.lower():
	    		    dict['switch'].append(hostname)
        		elif 'pdu' in device.lower():
	    		    dict['pdu'].append(hostname)
        		elif 'cisco' in device.lower():
	    		    dict['cisco'].append(hostname)
        		elif 'bda server' in device.lower():
	    		    dict['dom0'].append(hostname)
	    		    dict['ilom'].append(hostname + '-ilom')

	                rowNum = rowNum + 1

        #bda_list = []
        #ilom_list = []
        #switch_list = []
        #pdu_list = []
        #cisco_list = []
        
        #for device_val, hostname_val in zip(devices, hostnames):
        #    if 'switch' in device_val.lower():
        #        switch_list.append(hostname_val)
        #    elif 'pdu' in device_val.lower():
        #        pdu_list.append(hostname_val)
        #    elif 'bdcs' in hostname_val:
        #        cisco_list.append(hostname_val)
        #    elif 'bda server' in device_val.lower():
        #        bda_list.append(hostname_val)
        #        ilom_list.append(hostname_val + '-ilom')

        
        def iterate_list(a_list):
            for host in a_list:
                ip_address = self.__get_ip(host)
                self.__append_to_file('%s %s %s' %(host, self.internal_dns_domain, ip_address))
                
        #iterate_list(bda_list)
        iterate_list(dict['dom0'])
        #iterate_list(ilom_list)
        iterate_list(dict['ilom'])
        #iterate_list(switch_list)
        iterate_list(dict['switch'])
        #iterate_list(pdu_list)
        iterate_list(dict['pdu'])
        #iterate_list(cisco_list)
        iterate_list(dict['cisco'])
    
    
#    def __get_nw_rows(self, col, rowNum):
#	nullcount = 0
#        while True:
#	    if self.sheet.cell(row=rowNum, column=col).value == None:
#		nullcount = nullcount + 1    
#
#	        if nullcount > 2:
#		    break
#
#	        rowNum = rowNum + 1
#	        continue 
#
#	    value = self.sheet.cell(row=rowNum, column=col).value
#
#	    if re.match(r'\d+\.\d+\.\d+\.\d+/\d+', value):
#	        rowNum = rowNum + 1
#	        nullcount = 0	
#	    else:
#		break
#
#	return rowNum


    def get_nw_info(self):
	#
	# function to grab n/w details from RE sheet
	#
	# args: none
	#
	# dependencies: 
	#	variables - subnet.col, subnet.row
	#	
	#	functions - ipcalc, __append_to_file()
	#
	# returns: none
	#
	# side effects: updates rack text files
	#

        nullcount = 0
	nwinfo = {}
	interface = 1
	col = self.subnet.col 
	rowNum = self.subnet.row+1

        while True:
            if self.sheet.cell(row=rowNum, column=col).value == None:
                nullcount = nullcount + 1

                if nullcount > 2:
                    break

                rowNum = rowNum + 1
                continue

            subnet = self.sheet.cell(row=rowNum, column=col).value
            if re.match(r'\d+\.\d+\.\d+\.\d+/\d+', subnet):
		nwinfo[interface] = {}
		nwinfo[interface]["subnet"] = subnet
                #rowNum = rowNum + 1
                nullcount = 0
            else:
                break

	    vlan = self.sheet.cell(row=rowNum, column=col-3).value
	    vlan = unicode(vlan)
	    if re.match(r'\d+', vlan):
		nwinfo[interface]["vlan"] = vlan
	    else:
		raise ValueError("Given VLAN number is invalid for subnet: %s" %(subnet))

	    category = self.sheet.cell(row=rowNum, column=col-4).value
       	    if re.match(r'\S+', category):
	        nwinfo[interface]["category"] = category
	    else:
		raise ValueError("Given interface category is invalid for subnet: %s" %(subnet))

            rowNum = rowNum + 1
	    interface = interface + 1

	#print nwinfo
	###############################################################

        #vlan_info_count = 10
	# get the last row number of the network details section in spreadsheet
        #vlan_info_count = self.__get_nw_rows(self.subnet.col, self.subnet.row+1)

        #interface_category = self.__get_rows(self.interface.col, self.interface.row+1, self.interface.row + vlan_info_count)
        #subnet_details = self.__get_rows(self.subnet.col, self.subnet.row+1, self.subnet.row + vlan_info_count)
        #vlans = self.__get_rows(self.vlan.col, self.vlan.row+1, self.vlan.row + vlan_info_count) 
   
        class NwInfo(object):
            def __init__(self, ipcalc_nw_obj=None, pkey=None, vlan=None, connector=None):
                self.ipcalc_nw_obj = ipcalc_nw_obj
                self.pkey = pkey
                self.vlan = vlan
                self.connector = connector
                

        fe_list = []
        dom0_host = None
        connectors = 0
        #for i, interface_val in enumerate(interface_category):
	for i in nwinfo:
            #nw = ipcalc.Network(subnet_details[i])
	    nw = ipcalc.Network(nwinfo[i]["subnet"])
            #vlan = str(vlans[i])
	    vlan = nwinfo[i]["vlan"]
            pkey_val = '3a%s' %(vlan[-2:])
            vlan_val = 'some-emea-name-v%s' %(vlan)
	    interface_val = nwinfo[i]["category"]
            
            ###if interface_val.lower() == 'FE'.lower():
	    if 'FE'.lower() in interface_val.lower(): 
                nw_info_obj = NwInfo(ipcalc_nw_obj=nw, pkey=pkey_val, vlan=vlan_val, connector=connectors+1)
                fe_list.append(nw_info_obj)
                connectors = connectors + 1

            if vlan == '300':
                dom0_host = NwInfo(ipcalc_nw_obj=nw, vlan=vlan_val)

        for fe in fe_list:
            nw = fe.ipcalc_nw_obj
            self.__append_to_file('\n##############################################################################')
            self.__append_to_file('# Network:   \t%s' %(nw.network()))
            self.__append_to_file('# Netmask:   \t%s' %(nw.netmask()))
            self.__append_to_file('# Broadcast: \t%s' %(nw.broadcast()))
            self.__append_to_file('# Gateway:   \t%s' %(nw.host_first()))
            self.__append_to_file('# VLAN:      \t%s' %(fe.vlan))
            self.__append_to_file('# PKEY:      \t%s' %(fe.pkey))
            self.__append_to_file('# Connectors:\t%d' %(fe.connector))

        nw = dom0_host.ipcalc_nw_obj
        self.__append_to_file('\n##############################################################################')
        self.__append_to_file('# Network:   \t%s' %(nw.network()))
        self.__append_to_file('# Netmask:   \t%s' %(nw.netmask()))
        self.__append_to_file('# Broadcast: \t%s' %(nw.broadcast()))
        self.__append_to_file('# Gateway:   \t%s' %(nw.host_first()))
        self.__append_to_file('# VLAN:      \t%s' %(dom0_host.vlan))


def findBdcsId(wb, pod_name, region_name, cage_name, first_sheet):
    if not first_sheet:
        for sheet_name in wb.sheetnames:
            if "BDA" in sheet_name:
                first_sheet = sheet_name.split(" ")[0]
                break;
    
    return 'BDCS_%s_%s_%s_%s' %(pod_name, region_name, cage_name, first_sheet)
    
def get_json(json_file):
    data = json.loads('{}')
    data_file = open(json_file, 'r')

    try:
        data = json.load(data_file)
    finally:
        data_file.close()

    return data

if __name__ == '__main__':
    parser = optparse.OptionParser()
    parser.add_option('--xls_file', dest='xls_file', help='Enter xls file', type='string')
    parser.add_option('--json_file', dest='json_file', help='Enter json file', type='string')
    #parser.add_option('--first_sheet', dest='first_sheet', help='Enter total racks', type='string')
    parser.add_option('--first_dom0', dest='first_dom0', help='Enter first Dom0 hostname', type='string')
    #parser.add_option('--total_racks', dest='total_racks', help='Enter total racks', type='int')
    (options, args) = parser.parse_args()
    
    json_file_path = None
    if not options.json_file:
        parser.error('Json file not given')
    json_file_path = options.json_file
        
    xls_file_path = None
    if not options.xls_file:
        parser.error('Excel file not given')
    xls_file_path = options.xls_file
    
    #first_sheet = None
    #if options.first_sheet:
    #    first_sheet = options.first_sheet
    
    if not options.first_dom0:
        parser.error('First dom0 hostname not given')
    first_dom0 = options.first_dom0

    #total_sheets = None
    #if not options.total_racks:
    #    total_sheets = 5
    #else:
    #    total_sheets = int(options.total_racks)
        
    #sheet_numbers = [int(first_sheet) + i for i in range(0, total_sheets)]
    #print sheet_numbers
    
    wb = openpyxl.load_workbook(filename=xls_file_path, read_only=True)
    pod_dict = get_json(json_file_path)
    
    file_name = os.path.splitext(xls_file_path)[0]
    file_name_list = file_name.split('-')
    pod_dict['region_name'] = file_name_list[2]
    pod_dict['cage_name'] = (file_name_list[3].split('_'))[1]
    #pod_dict['bdcs_zone_id'] = findBdcsId(wb, pod_dict['pod_name'], pod_dict['region_name'], pod_dict['cage_name'], first_sheet)


    # Prepare a list of first dom0 of each sheet
    print "\nINFO: Preparing list of first dom0 hostnames for each rack\n"
    dom0_list = [first_dom0]
    prev_dom0 = first_dom0

    for i in range(1,5):
	string = re.sub(r'\d', "", prev_dom0)
	num = re.sub(r'\D', "", prev_dom0)

	num = int(num) + 18
	#if len(str(num)) < 4:
	while(len(str(num)) < 4):
	    num = "0" + str(num)

	prev_dom0 = string + str(num)
	dom0_list.append(prev_dom0)

    print "\nINFO: Searching for below dom0's in given spreadsheet\n"
    print dom0_list

    # Prepare the list of sheet names where above dom0's exists
    sheetnames_to_process = []

    found = 0
    for dom0 in dom0_list:
	for sheet in wb.sheetnames:
	    if "Overview" in sheet:
		continue    
	    if "AM3-Rack Elevation" in sheet:
		continue    

  	    for row in wb[sheet].rows:
		for cell in row:
		    if cell.value == dom0:
			sheetnames_to_process.append(sheet)
			found = 1

			if dom0 == first_dom0:
			    pod_dict['bdcs_zone_id'] = 'BDCS_%s_%s_%s_%s' %(pod_dict['pod_name'], pod_dict['region_name'], pod_dict['cage_name'], sheet.split(" ")[0])

			break

		if found == 1:
		    break

	    if found == 1:
	        found = 0
		break

    if len(sheetnames_to_process) < 5:
        raise ValueError("One or more dom0 hostnames are not found in given RE sheet")
 
    if len(sheetnames_to_process) > 5:
        raise ValueError("dom0 hostnames are found in multiple places of given RE sheet")

    txt_file_prefix = 'BDCS-%s-%s' %(pod_dict['region_name'], pod_dict['cage_name'])

#    sheet_num = 1
#    for sheet_name in wb.sheetnames:
#        if "BDA" in sheet_name:
#            for x in sheet_numbers:
#                if str(x) in sheet_name:
#                    txt_file_name = '%s-%s' %(txt_file_prefix, sheet_name.split(" ")[0])
#                    xl_sheet = XLSheet(wb[sheet_name], pod_dict, txt_file_name)
#                    xl_sheet.create_text_file()
#                    sheet_num = sheet_num + 1
            
#        if sheet_num > total_sheets:
#            break

    for sheet_name in sheetnames_to_process:
	txt_file_name = '%s-%s' %(txt_file_prefix, sheet_name.split(" ")[0])
	xl_sheet = XLSheet(wb[sheet_name], pod_dict, txt_file_name)
	xl_sheet.create_text_file()

