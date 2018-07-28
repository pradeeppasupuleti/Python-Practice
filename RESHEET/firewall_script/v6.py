import openpyxl
import re, sys, logging, time, warnings, argparse

parser = argparse.ArgumentParser(add_help=True)
parser.add_argument("--re_sheet")
parser.add_argument("--template_file")
parser.add_argument("--log_file")
args = parser.parse_args()

INPUT_XLSX = args.re_sheet
TEMPLATE_FILE = args.template_file
LOG_FILE_PATH = args.log_file


if LOG_FILE_PATH == None or INPUT_XLSX == None or TEMPLATE_FILE == None:
    parser.print_usage()
    sys.exit(2)
    

warnings.simplefilter("ignore")
day=(time.strftime("%d-%m-%Y"))
logging.basicConfig(filename=LOG_FILE_PATH, level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def getSubnetIPs(book = "", name = ""):
    xl_sheet = book[name]
    logging.info('Processing BDA sheet - {0}'.format(name))
    req_col_name_01 = "INTERFACE CATEGORY"
    req_col_name_02 = "Subnet Details"
    req_columns_found = False
    req_column_index_01 = 0
    req_column_index_02 = 0
    sheet_ips_list = []
    m_row = xl_sheet.max_row
    m_col = xl_sheet.max_column
    
    for row_idx in range(1, m_row):
        row_data = []
        for col_idx in range(1, m_col):
            cell_obj = xl_sheet.cell(row = row_idx, column = col_idx)
            if cell_obj:
                row_data.append(str(cell_obj.value))
            else:
                row_data.append(str("NA"))
                                
        if (req_col_name_01 in row_data) and (req_col_name_02 in row_data):    # Checking for the required value in the 
            req_column_index_01 = row_data.index(req_col_name_01)
            req_column_index_02 = row_data.index(req_col_name_02)
            req_columns_found = True
            continue
        else:
            pass
        
        if req_columns_found and req_column_index_01 and req_column_index_02:
            if row_data[req_column_index_02] and row_data[req_column_index_01]:
                if row_data[req_column_index_01].lower().startswith('customer'):
                    sheet_ips_list.append(row_data[req_column_index_02])
    return sheet_ips_list



def get_fe_be_mgmt_IPs(book = "", name = ""):
#     xl_sheet = book.get_sheet_by_name(name)
    xl_sheet = book[name]
    logging.info('Processing exdata sheet - {0}'.format(name))
    req_col_name_01 = "INTERFACE CATEGORY"
    req_col_name_02 = "Subnet Details"
    req_columns_found = False
    req_column_index_01 = 1
    req_column_index_02 = 1
    fe_ips_list = []
    be_ips_list = []
    mgmt_ips_list = []
    bdcs_mgmt_ip = []
    m_row = xl_sheet.max_row
    m_col = xl_sheet.max_column
    
    for row_idx in range(1, m_row):    # Iterate through rows
        row_data = []
        for col_idx in range(1, m_col):  # Iterate through columns
            cell_obj = xl_sheet.cell(row = row_idx, column = col_idx)  # Get cell object by row, col
            if cell_obj:
                row_data.append(str(cell_obj.value))
            else:
                row_data.append(str("NA"))
        if (req_col_name_01 in row_data) and (req_col_name_02 in row_data):
            req_column_index_01 = row_data.index(req_col_name_01)
            req_column_index_02 = row_data.index(req_col_name_02)
            req_columns_found = True
            continue
        else:
            pass
            
        if req_columns_found and req_column_index_01 and req_column_index_02:
            if row_data[req_column_index_02] and row_data[req_column_index_01]:
                if row_data[req_column_index_01].lower().startswith('customer') and row_data[req_column_index_01].lower().endswith('fe'):
                    fe_ips_list.append(row_data[req_column_index_02])
                elif row_data[req_column_index_01].lower().startswith('customer') and row_data[req_column_index_01].lower().endswith('be'):
                    be_ips_list.append(row_data[req_column_index_02])
                elif row_data[req_column_index_01].lower().startswith('customer') and row_data[req_column_index_01].lower().endswith('mgmt'):
                    mgmt_ips_list.append(row_data[req_column_index_02])
                elif row_data[req_column_index_01] == "BDCS Management":
                    bdcs_mgmt_ip.append(row_data[req_column_index_02])
                                   
    return fe_ips_list, be_ips_list, mgmt_ips_list, bdcs_mgmt_ip
    
def process_all_sheets():
    all_info = []
    for sheet in sheet_name:          
        if re.match(r'\w\d\w\d+ BDA$', sheet):
            try:    
                sheet_data = getSubnetIPs(book = wb, name = sheet)        
                all_info += sheet_data
            except Exception as e:
                logging.critical("Exception > process_all_sheets() for BDA/BDCS -> {0}".format(e))
    fe_info = []
    be_info = []
    mgmt_info = []
    bdcs_mgmt_info = []
    for sheet in sheet_name:
        if sheet.lower().endswith("exdata"):
            try:
                fe_data, be_data, mgmt_data, bdcs_mgmt = get_fe_be_mgmt_IPs(book = wb, name = sheet)
                fe_info += fe_data
                be_info += be_data
                mgmt_info += mgmt_data
                bdcs_mgmt_info += bdcs_mgmt
            except Exception as e:
                logging.critical("Exception > process_all_sheets() for exdata -> {0}".format(e))
    if all_info and bdcs_mgmt_info and mgmt_info and be_info and fe_info:
        logging.info("Generating output xlsx ...")
        wb1 = openpyxl.load_workbook(TEMPLATE_FILE)
        ws = wb1.active
        ws["D2"] = ("\n".join(all_info))
        ws["D3"] = ("\n".join(fe_info))
        ws["D4"] = ("\n".join(be_info))
        ws["D5"] = ("\n".join(mgmt_info))
        ws["D6"] = ("\n".join(all_info))
        ws["D7"] = ("\n".join(all_info))
        ws["A8"] = ("\n".join(all_info))
        ws["D9"] = ("\n".join(bdcs_mgmt_info))
        ws["D10"] = ("\n".join(fe_info))
        wb1.save(day + '_' + TEMPLATE_FILE)
        logging.info(day + '_' + TEMPLATE_FILE + " " + "sheet Generated Successfully...")
    else:
        logging.critical("Unable to generate sheet dont have all the data .....")
        sys.exit()
                
if INPUT_XLSX.endswith("xlsx") and TEMPLATE_FILE.endswith("xls"):
    try:
        wb = openpyxl.load_workbook(INPUT_XLSX)
        sheet_name = wb.sheetnames
        logging.info('--------------------------------------------------------------------')
        logging.info("Processing  started with the given input file" + "---->" + "'" + str(INPUT_XLSX) + "'")
        logging.info("List of input sheets - {0}".format(sheet_name))
        logging.info("Number of sheets found in the file" + ":" + str(len(sheet_name)))
        
    except Exception as k:
        logging.critical("exception occurred - {0}".format(k))
        sys.exit()
else:
    logging.info("Execution format -> # python parse_xlsx.py --re_sheet <xlsx/xls> --template_file <xlsx/xls> --log_file <value>")
    logging.info("Exiting the script input file format is not correct...")
    sys.exit()
     
#Executing the main Function    
if __name__ == '__main__':
    try:
        process_all_sheets()
        logging.info('--------------------------------------------------------------------')
    except Exception as emsg:
        logging.critical("exception occurred - {0}".format(emsg))     



