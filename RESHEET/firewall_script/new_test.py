import openpyxl
import re
import sys
import logging
import time
day=(time.strftime("%d-%m-%Y"))
logging.basicConfig(filename='new_firewall.log', level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
import warnings
warnings.simplefilter("ignore")


def getSubnetIPs(book = "", name = ""):
    xl_sheet = book[name]
    logging.info('Processing BDA sheet - {0}'.format(name))
    req_col_name_01 = "INTERFACE CATEGORY"
    req_col_name_02 = "Subnet Details"
    req_columns_found = False
    req_column_index_01 = 0
    req_column_index_02 = 0
    sheet_ips_list = []
    m_row = xl_sheet.max_row
    m_col = xl_sheet.max_column
    
    for row_idx in range(1, m_row):
        row_data = []
        for col_idx in range(1, m_col):
            cell_obj = xl_sheet.cell(row = row_idx, column = col_idx)
            if cell_obj:
                row_data.append(str(cell_obj.value))
            else:
                row_data.append(str("NA"))
                                
        if (req_col_name_01 in row_data) and (req_col_name_02 in row_data):    # Checking for the required value in the 
            #print("Found.. in row - {0}".format(row_idx))
            req_column_index_01 = row_data.index(req_col_name_01)
            req_column_index_02 = row_data.index(req_col_name_02)
            req_columns_found = True
            continue
        else:
            pass
        
        if req_columns_found and req_column_index_01 and req_column_index_02:
            if row_data[req_column_index_02] and row_data[req_column_index_01]:
                if row_data[req_column_index_01].lower().startswith('customer'):
                    sheet_ips_list.append(row_data[req_column_index_02])
    return sheet_ips_list



def get_fe_be_mgmt_IPs(book = "", name = ""):
    xl_sheet = book[name]
    logging.info('Processing exdata sheet - {0}'.format(name))
    req_col_name_01 = "INTERFACE CATEGORY"
    req_col_name_02 = "Subnet Details"
    req_columns_found = False
    req_column_index_01 = 1
    req_column_index_02 = 1
    fe_ips_list = []
    be_ips_list = []
    mgmt_ips_list = []
    bdcs_mgmt_ip = []
    m_row = xl_sheet.max_row
    m_col = xl_sheet.max_column
    
    for row_idx in range(1, m_row):    # Iterate through rows
        #print ('Row: %s' % row_idx)
        row_data = []
        for col_idx in range(1, m_col):  # Iterate through columns
            cell_obj = xl_sheet.cell(row = row_idx, column = col_idx)  # Get cell object by row, col
#             pprint(cell_obj)
            if cell_obj:
                row_data.append(str(cell_obj.value))
#                 pprint(row_data)
            else:
                row_data.append(str("NA"))
        if (req_col_name_01 in row_data) and (req_col_name_02 in row_data):
            req_column_index_01 = row_data.index(req_col_name_01)
            req_column_index_02 = row_data.index(req_col_name_02)
            req_columns_found = True
            continue
            #process further
        else:
            pass
            
        if req_columns_found and req_column_index_01 and req_column_index_02:
            if row_data[req_column_index_02] and row_data[req_column_index_01]:
                if row_data[req_column_index_01].lower().startswith('customer') and row_data[req_column_index_01].lower().endswith('fe'):
                    fe_ips_list.append(row_data[req_column_index_02])
                elif row_data[req_column_index_01].lower().startswith('customer') and row_data[req_column_index_01].lower().endswith('be'):
                    be_ips_list.append(row_data[req_column_index_02])
                elif row_data[req_column_index_01].lower().startswith('customer') and row_data[req_column_index_01].lower().endswith('mgmt'):
                    mgmt_ips_list.append(row_data[req_column_index_02])
                elif row_data[req_column_index_01] == "BDCS Management":
                    bdcs_mgmt_ip.append(row_data[req_column_index_02])
                                   
    return fe_ips_list, be_ips_list, mgmt_ips_list, bdcs_mgmt_ip
    
    
# book = openpyxl.load_workbook('cloud01.xlsx')
# sheet_name = book.sheetnames

def process_all_sheets():
    all_info = []
    for sheet in sheet_name:          
        if re.match(r'\w\d\w\d+ BDA$', sheet):
            try:    
                sheet_data = getSubnetIPs(book = book, name = sheet)        
                all_info += sheet_data
            except Exception as e:
                logging.critical("Exception > process_all_sheets() for BDA/BDCS -> {0}".format(e))
    fe_info = []
    be_info = []
    mgmt_info = []
    bdcs_mgmt_info = []
    for sheet in sheet_name:
        if sheet.lower().endswith("exdata"):
            try:
                fe_data, be_data, mgmt_data, bdcs_mgmt = get_fe_be_mgmt_IPs(book = book, name = sheet)
                fe_info += fe_data
                be_info += be_data
                mgmt_info += mgmt_data
                bdcs_mgmt_info += bdcs_mgmt
            except Exception as e:
                logging.critical("Exception > process_all_sheets() for exdata -> {0}".format(e))
    if all_info and bdcs_mgmt_info and mgmt_info and be_info and fe_info:
        logging.info("Generating output xlsx ...")
#             style2=xlwt.easyxf('align: wrap yes; font: name Calibri, height 220')
        wb1 = openpyxl.load_workbook(sys.argv[2])
        ws = wb1.active
#             sheet1 = wb1.get_sheet(0)
        ws["D2"] = ("\n".join(all_info))
        ws["D3"] = ("\n".join(fe_info))
        ws["D4"] = ("\n".join(be_info))
        ws["D5"] = ("\n".join(mgmt_info))
        ws["D6"] = ("\n".join(all_info))
        ws["D7"] = ("\n".join(all_info))
        ws["A8"] = ("\n".join(all_info))
        ws["D9"] = ("\n".join(bdcs_mgmt_info))
        ws["D10"] = ("\n".join(fe_info))
        wb1.save(day + '_' + sys.argv[2])
        logging.info(day + '_' + sys.argv[2] + " " + "sheet Generated Successfully...")
    else:
        logging.critical("Unable to generate sheet dont have all the data .....")
        sys.exit()
                


# # if len(sys.argv) == 3 and ( str(sys.argv[1]).endswith("xlsx") or str(sys.argv[1]).endswith("xls") ) and str(sys.argv[2]).endswith("xls"):
if len(sys.argv) == 3 and ( str(sys.argv[1]).endswith("xlsx") or str(sys.argv[1]).endswith("xls")):
    try:
        book = openpyxl.load_workbook(sys.argv[1])
        sheet_name = book.sheetnames
        logging.info('--------------------------------------------------------------------')
        logging.info("Processing  started with the given input file" + "---->" + "'" + str(sys.argv[1]) + "'")
        logging.info("List of input sheets - {0}".format(sheet_name))
        logging.info("Number of sheets found in the file" + ":" + str(len(sheet_name)))
        
    except Exception as k:
        logging.critical("exception occurred - {0}".format(k))
        sys.exit()
else:
    logging.info("Execution format -> # python parse_xlsx.py <xlsx / xls file>")
    logging.info("Exiting the script input file format is not correct...")
    sys.exit()
     
#Executing the main Function    
if __name__ == '__main__':
    try:
        process_all_sheets()
        logging.info('--------------------------------------------------------------------')
    except Exception as emsg:
        logging.critical("exception occurred - {0}".format(emsg))     

